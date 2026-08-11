"""导出接口测试缺陷报告到 Excel"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 输出路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(BASE_DIR, "docs", "接口测试缺陷报告.xlsx")

# 缺陷数据
defects = [
    {
        "id": "BUG-001",
        "title": "业主越权访问用户管理接口",
        "severity": "P0-致命",
        "priority": "立即修复",
        "module": "用户管理/安全",
        "api": "GET /user/page",
        "test_case": "test_user_api.py::TestUserPageAPI::test_page_query_user_role",
        "description": "业主使用 USER 类型 Token 可以成功调用 GET /user/page 接口，获取到全部用户列表数据。系统未对 /user/* 接口进行角色权限校验。",
        "steps": "1. 业主 user1/123456 登录获取 Token\n2. 使用该 Token 调用 GET /user/page?pageNum=1&pageSize=10\n3. 观察响应结果",
        "expected": "HTTP 401，返回无权限错误提示",
        "actual": "HTTP 200，返回全部用户列表数据",
        "impact": "用户隐私泄露（用户名、手机号等），可能导致进一步越权操作",
        "suggestion": "在 LoginInterceptor 或 UserController 中增加角色校验，仅允许 ADMIN 类型用户访问 /user/* 接口",
    },
    {
        "id": "BUG-002",
        "title": "业主越权处理报修",
        "severity": "P0-致命",
        "priority": "立即修复",
        "module": "报修管理/安全",
        "api": "POST /repair/handle/{id}",
        "test_case": "test_repair_api.py::TestRepairHandleAPI::test_handle_repair_user_role",
        "description": "业主使用 USER 类型 Token 可以成功调用 POST /repair/handle/{id} 接口，将报修状态修改为已处理。系统未对该接口进行角色权限校验。",
        "steps": "1. 业主 user1/123456 登录获取 Token\n2. 使用该 Token 调用 POST /repair/handle/1\n3. 观察响应结果",
        "expected": "HTTP 401，返回无权限错误提示",
        "actual": "HTTP 200，报修状态被成功修改",
        "impact": "报修流程被恶意篡改，管理员无法正常跟踪报修处理状态",
        "suggestion": "在 RepairController.handle() 方法中增加角色校验，仅允许 ADMIN 类型用户调用",
    },
    {
        "id": "BUG-003",
        "title": "分页参数 pageNum=0 导致服务器 500 错误",
        "severity": "P1-严重",
        "priority": "上线前修复",
        "module": "用户管理",
        "api": "GET /user/page",
        "test_case": "test_user_api.py::TestUserPageAPI::test_page_query_boundary_page_zero",
        "description": "当分页查询接口的 pageNum 参数传值为 0 时，后端抛出未处理异常，返回 HTTP 500 错误。",
        "steps": "1. 管理员登录获取 Token\n2. 调用 GET /user/page?pageNum=0&pageSize=10\n3. 观察响应结果",
        "expected": "HTTP 200，返回第一页数据或参数错误提示",
        "actual": "HTTP 500，服务器内部错误",
        "impact": "影响所有分页查询接口，前端传入异常参数时页面崩溃",
        "suggestion": "在 Service 层对 pageNum 进行校验，当 pageNum < 1 时自动设置为 1",
    },
    {
        "id": "BUG-004",
        "title": "分页参数 pageNum=-1 导致服务器 500 错误",
        "severity": "P1-严重",
        "priority": "上线前修复",
        "module": "用户管理",
        "api": "GET /user/page",
        "test_case": "test_user_api.py::TestUserPageAPI::test_page_query_boundary_negative_page",
        "description": "当分页查询接口的 pageNum 参数传值为负数（如 -1）时，后端抛出未处理异常，返回 HTTP 500 错误。",
        "steps": "1. 管理员登录获取 Token\n2. 调用 GET /user/page?pageNum=-1&pageSize=10\n3. 观察响应结果",
        "expected": "HTTP 200，返回第一页数据或参数错误提示",
        "actual": "HTTP 500，服务器内部错误",
        "impact": "同 BUG-003",
        "suggestion": "同 BUG-003，统一对分页参数进行下限校验",
    },
    {
        "id": "BUG-005",
        "title": "新增用户接口返回 500 错误",
        "severity": "P1-严重",
        "priority": "上线前修复",
        "module": "用户管理",
        "api": "POST /user/add",
        "test_case": "test_user_api.py::TestUserAddAPI::test_add_user_success",
        "description": "调用 POST /user/add 新增用户时，无论传入合法还是非法参数，后端均返回 HTTP 500 错误。",
        "steps": "1. 管理员登录获取 Token\n2. 调用 POST /user/add，Body: {username:testuser,nickname:测试用户,password:123456}\n3. 观察响应结果",
        "expected": "HTTP 200，返回 {code:200,msg:操作成功}",
        "actual": "HTTP 500，服务器内部错误",
        "impact": "管理员无法通过接口新增用户，前端新增用户功能不可用",
        "suggestion": "检查 UserServiceImpl.add() 方法，排查数据库约束、必填字段校验、异常处理逻辑",
    },
    {
        "id": "BUG-006",
        "title": "创建物业费账单接口返回 500 错误",
        "severity": "P1-严重",
        "priority": "上线前修复",
        "module": "费用管理",
        "api": "POST /propertyFee/add",
        "test_case": "test_fee_api.py::TestPropertyFeeAddAPI::test_add_fee_success",
        "description": "调用 POST /propertyFee/add 创建物业费账单时，后端返回 HTTP 500 错误。各种参数场景均返回 500。",
        "steps": "1. 管理员登录获取 Token\n2. 调用 POST /propertyFee/add，Body: {houseId:1,fee:200.00}\n3. 观察响应结果",
        "expected": "HTTP 200，返回 {code:200,msg:操作成功}",
        "actual": "HTTP 500，服务器内部错误",
        "impact": "管理员无法创建物业费账单，物业费管理功能完全不可用",
        "suggestion": "检查 PropertyFeeServiceImpl.add() 方法，排查 houseId 关联查询、userId 获取逻辑、数据库插入操作",
    },
    {
        "id": "BUG-007",
        "title": "提交报修接口返回 500 错误",
        "severity": "P1-严重",
        "priority": "上线前修复",
        "module": "报修管理",
        "api": "POST /repair/add",
        "test_case": "test_repair_api.py::TestRepairAddAPI::test_add_repair_success",
        "description": "调用 POST /repair/add 提交报修时，后端返回 HTTP 500 错误。各种参数场景均返回 500。",
        "steps": "1. 业主登录获取 Token\n2. 调用 POST /repair/add，Body: {houseId:1,type:水电维修,content:水龙头漏水}\n3. 观察响应结果",
        "expected": "HTTP 200，返回 {code:200,msg:操作成功}",
        "actual": "HTTP 500，服务器内部错误",
        "impact": "业主无法提交报修，报修管理功能不可用",
        "suggestion": "检查 RepairServiceImpl.add() 方法，排查 houseId 关联查询、userId 设置逻辑、数据库插入操作",
    },
    {
        "id": "BUG-008",
        "title": "XSS 内容导致服务器 500 错误",
        "severity": "P1-严重",
        "priority": "上线前修复",
        "module": "报修管理/安全",
        "api": "POST /repair/add",
        "test_case": "test_repair_api.py::TestRepairAddAPI::test_add_repair_xss_content",
        "description": "当报修内容包含 XSS 攻击载荷（如 <script>alert('xss')</script>）时，后端返回 HTTP 500 错误。",
        "steps": "1. 业主登录获取 Token\n2. 调用 POST /repair/add，Body: {houseId:1,type:水电维修,content:<script>alert('xss')</script>}\n3. 观察响应结果",
        "expected": "HTTP 200，内容被转义后存储，或返回参数错误提示",
        "actual": "HTTP 500，服务器内部错误",
        "impact": "用户输入特殊字符时功能不可用，可能存在安全风险",
        "suggestion": "1. 排查 500 错误根因 2. 对输入内容进行 HTML 转义 3. 增加输入内容长度限制",
    },
    {
        "id": "BUG-009",
        "title": "禁用用户仍可登录",
        "severity": "P1-严重",
        "priority": "上线前修复",
        "module": "用户认证",
        "api": "POST /common/login",
        "test_case": "test_auth_api.py::TestLoginAPI::test_login_disabled_user",
        "description": "状态为禁用的用户（如 user3）仍可通过登录接口正常登录并获取 Token，系统未在登录时校验用户状态。",
        "steps": "1. 调用 POST /common/login，Body: {username:user3,password:123456,type:USER}\n2. 观察响应结果",
        "expected": "登录失败，返回账号已禁用提示",
        "actual": "登录成功，返回有效的 JWT Token",
        "impact": "被禁用的用户仍可操作系统，管理员禁用用户操作无效",
        "suggestion": "在 CommonService.login() 方法中增加用户状态校验，当 status 为禁用时抛出异常",
    },
    {
        "id": "BUG-010",
        "title": "负数充值未校验",
        "severity": "P1-严重",
        "priority": "上线前修复",
        "module": "用户管理",
        "api": "POST /user/topUp/{amount}",
        "test_case": "test_user_api.py::TestUserTopUpAPI::test_top_up_negative",
        "description": "业主调用充值接口传入负数金额时（如 -100），系统未拒绝请求，反而成功扣减了业主余额。",
        "steps": "1. 业主登录获取 Token\n2. 调用 POST /user/topUp/-100\n3. 观察响应结果和余额变化",
        "expected": "充值失败，返回充值金额必须大于0提示，余额不变",
        "actual": "充值成功，余额被扣减 100",
        "impact": "余额数据被篡改，可能影响支付逻辑",
        "suggestion": "在 UserController.topUp() 或 UserService.topUp() 方法中增加金额校验，当 amount <= 0 时抛出异常",
    },
    {
        "id": "BUG-011",
        "title": "处理不存在的报修 ID 返回成功",
        "severity": "P2-一般",
        "priority": "建议修复",
        "module": "报修管理",
        "api": "POST /repair/handle/{id}",
        "test_case": "test_repair_api.py::TestRepairHandleAPI::test_handle_nonexistent_repair",
        "description": "调用 POST /repair/handle/99999 处理一个不存在的报修记录时，系统返回成功。",
        "steps": "1. 管理员登录获取 Token\n2. 调用 POST /repair/handle/99999\n3. 观察响应结果",
        "expected": "返回失败，提示报修记录不存在",
        "actual": "返回成功",
        "impact": "误导管理员操作，可能产生脏数据",
        "suggestion": "在 RepairServiceImpl.handle() 方法中，先查询报修记录是否存在，不存在时抛出异常",
    },
    {
        "id": "BUG-012",
        "title": "业主越权创建物业费账单返回 500",
        "severity": "P2-一般",
        "priority": "建议修复",
        "module": "费用管理/安全",
        "api": "POST /propertyFee/add、POST /common/resetPassword",
        "test_case": "test_fee_api.py::TestPropertyFeeAddAPI::test_add_fee_user_role",
        "description": "业主使用 USER 类型 Token 调用管理员接口时，后端返回 HTTP 500 而非 HTTP 401。系统未进行显式角色权限校验。",
        "steps": "1. 业主登录获取 Token\n2. 调用 POST /propertyFee/add，Body: {houseId:1,fee:100.00}\n3. 观察响应结果",
        "expected": "HTTP 401，返回无权限错误提示",
        "actual": "HTTP 500，服务器内部错误",
        "impact": "错误信息不友好，暴露系统内部异常，权限校验依赖代码副作用",
        "suggestion": "在相关 Controller 方法上增加角色校验注解，对非管理员用户直接返回 401",
    },
]


def create_excel():
    wb = Workbook()

    # ===== Sheet 1: 缺陷汇总 =====
    ws1 = wb.active
    ws1.title = "缺陷汇总"

    # 标题行
    headers = [
        "缺陷编号", "缺陷标题", "严重程度", "优先级", "所属模块",
        "关联接口", "关联用例", "缺陷描述", "复现步骤",
        "预期结果", "实际结果", "影响范围", "修复建议"
    ]

    # 样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell_font = Font(name="微软雅黑", size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 严重程度颜色
    severity_colors = {
        "P0-致命": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "P1-严重": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "P2-一般": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    }
    severity_fonts = {
        "P0-致命": Font(name="微软雅黑", size=10, bold=True, color="FFFFFF"),
        "P1-严重": Font(name="微软雅黑", size=10, bold=True, color="FFFFFF"),
        "P2-一般": Font(name="微软雅黑", size=10, bold=True, color="000000"),
    }

    # 写标题
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 写数据
    for row_idx, defect in enumerate(defects, 2):
        values = [
            defect["id"], defect["title"], defect["severity"], defect["priority"],
            defect["module"], defect["api"], defect["test_case"], defect["description"],
            defect["steps"], defect["expected"], defect["actual"], defect["impact"],
            defect["suggestion"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            # 严重程度列着色
            if col == 3 and value in severity_colors:
                cell.fill = severity_colors[value]
                cell.font = severity_fonts[value]

    # 列宽
    col_widths = [10, 25, 12, 12, 18, 25, 35, 40, 35, 25, 25, 30, 35]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = width

    # 行高
    ws1.row_dimensions[1].height = 30
    for row_idx in range(2, len(defects) + 2):
        ws1.row_dimensions[row_idx].height = 100

    # ===== Sheet 2: 统计汇总 =====
    ws2 = wb.create_sheet("统计汇总")

    # 统计数据
    stats = [
        ["统计项", "数量", "占比"],
        ["缺陷总数", 12, "100%"],
        ["P0-致命", 2, "16.7%"],
        ["P1-严重", 7, "58.3%"],
        ["P2-一般", 3, "25.0%"],
        ["", "", ""],
        ["按模块统计", "", ""],
        ["安全", 4, "33.3%"],
        ["用户管理", 4, "33.3%"],
        ["费用管理", 2, "16.7%"],
        ["报修管理", 3, "25.0%"],
        ["用户认证", 1, "8.3%"],
        ["", "", ""],
        ["按类型统计", "", ""],
        ["权限越界", 3, "25.0%"],
        ["输入校验缺失", 5, "41.7%"],
        ["服务器内部错误", 3, "25.0%"],
        ["业务逻辑缺陷", 1, "8.3%"],
    ]

    for row_idx, row_data in enumerate(stats, 1):
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if row_idx == 1 or (row_data[0] in ["按模块统计", "按类型统计"]):
                cell.font = Font(name="微软雅黑", size=11, bold=True)
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            else:
                cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10

    # ===== Sheet 3: 修复优先级 =====
    ws3 = wb.create_sheet("修复优先级")

    priority_data = [
        ["优先级", "缺陷编号", "缺陷标题", "建议修复时限"],
        ["立即修复", "BUG-001", "业主越权访问用户管理接口", "24小时内"],
        ["立即修复", "BUG-002", "业主越权处理报修", "24小时内"],
        ["上线前修复", "BUG-003", "分页参数 pageNum=0 导致 500", "3个工作日"],
        ["上线前修复", "BUG-004", "分页参数 pageNum=-1 导致 500", "3个工作日"],
        ["上线前修复", "BUG-005", "新增用户接口返回 500", "3个工作日"],
        ["上线前修复", "BUG-006", "创建物业费账单接口返回 500", "3个工作日"],
        ["上线前修复", "BUG-007", "提交报修接口返回 500", "3个工作日"],
        ["上线前修复", "BUG-008", "XSS 内容导致 500", "3个工作日"],
        ["上线前修复", "BUG-009", "禁用用户仍可登录", "3个工作日"],
        ["上线前修复", "BUG-010", "负数充值未校验", "3个工作日"],
        ["建议修复", "BUG-011", "处理不存在的报修 ID 返回成功", "5个工作日"],
        ["建议修复", "BUG-012", "业主越权返回 500", "5个工作日"],
    ]

    for row_idx, row_data in enumerate(priority_data, 1):
        for col, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = cell_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # 优先级着色
                if col == 1:
                    if value == "立即修复":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
                    elif value == "上线前修复":
                        cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
                        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
                    elif value == "建议修复":
                        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 35
    ws3.column_dimensions["D"].width = 15

    # 保存
    wb.save(OUTPUT_PATH)
    print(f"Excel 缺陷报告已生成: {OUTPUT_PATH}")


if __name__ == "__main__":
    create_excel()
